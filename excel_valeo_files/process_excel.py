import openpyxl
from openpyxl.styles import PatternFill

def process_excel(file_path, output_path):
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Create yellow fill pattern
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Initialize variables
    cached_row = None

    # Iterate through rows (starting from the second row, assuming the first row is a header)
    row_idx = 2
    while row_idx <= sheet.max_row:
        current_row = [cell.value for cell in sheet[row_idx]]
        next_row = [cell.value for cell in sheet[row_idx + 1]] if row_idx < sheet.max_row else None

        # Check if the third column contains 'head'
        if 'head' in str(current_row[2]):
            cached_row = current_row

        # Check the next row's second column for 'req'
        if next_row and 'req' in str(next_row[1]):
            # Compare the cached row with the current row
            if cached_row and cached_row != current_row:
                # Insert the cached row above the next row
                sheet.insert_rows(row_idx + 1)
                for col_idx, value in enumerate(cached_row, start=1):
                    cell = sheet.cell(row=row_idx + 1, column=col_idx, value=value)
                    cell.fill = yellow_fill  # Apply yellow highlight to the inserted row
                row_idx += 1  # Skip the inserted row

        row_idx += 1

    # Save the modified workbook
    wb.save(output_path)

# Example usage
input_file = "ff1.xlsx"  # Replace with your input file path
output_file = "output.xlsx"  # Replace with your desired output file path
process_excel(input_file, output_file)